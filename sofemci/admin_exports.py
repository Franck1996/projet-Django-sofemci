# P:\sofemci\sofemci\sofemci\admin_exports.py
from django.http import HttpResponse
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import datetime
from decimal import Decimal
from .admin_mixins import ExportPDFExcelMixin


class ProductionExportMixin(ExportPDFExcelMixin):
    """Mixin spécifique pour les modèles de production"""
    
    def export_pdf_view(self, request, period='today'):
        """Export PDF pour les modèles de production"""
        start_date, end_date = self.get_period_dates(period)
        
        # Filtrer les données
        queryset = self.get_export_queryset(start_date, end_date)
        
        # Créer le PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            spaceAfter=30,
            alignment=1
        )
        
        # Titre
        period_text = self.get_period_text(period)
        model_name = self.model._meta.verbose_name_plural
        title = Paragraph(f'Rapport {model_name} - {period_text}', title_style)
        elements.append(title)
        
        # Statistiques
        stats = self.get_production_stats(queryset)
        if stats:
            stats_text = self.format_stats_html(stats)
            stats_para = Paragraph(stats_text, styles['Normal'])
            elements.append(stats_para)
            elements.append(Spacer(1, 20))
        
        # Tableau des données
        data = self.get_pdf_table_data(queryset)
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(table)
        
        # Pied de page
        elements.append(Spacer(1, 30))
        footer = Paragraph(
            f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')} | "
            f"SOFEM-CI - Usine d'emballage",
            styles['Italic']
        )
        elements.append(footer)
        
        doc.build(elements)
        buffer.seek(0)
        
        # Retourner le PDF
        filename = f'{self.model._meta.model_name}_{period}_{datetime.date.today()}.pdf'
        return self.create_pdf_response(buffer, filename)
    
    def export_excel_view(self, request, period='today'):
        """Export Excel pour les modèles de production"""
        start_date, end_date = self.get_period_dates(period)
        queryset = self.get_export_queryset(start_date, end_date)
        
        # Créer le workbook Excel
        wb, ws = self.create_excel_workbook(queryset, period)
        
        # Sauvegarder dans le buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Retourner le fichier Excel
        filename = f'{self.model._meta.model_name}_{period}_{datetime.date.today()}.xlsx'
        return self.create_excel_response(buffer, filename)
    
    # Méthodes abstraites à implémenter dans chaque classe
    def get_export_queryset(self, start_date, end_date):
        """Retourne le queryset filtré pour l'export"""
        raise NotImplementedError
    
    def get_production_stats(self, queryset):
        """Calcule les statistiques pour le modèle"""
        raise NotImplementedError
    
    def get_pdf_table_data(self, queryset):
        """Retourne les données pour le tableau PDF"""
        raise NotImplementedError
    
    def create_excel_workbook(self, queryset, period):
        """Crée le workbook Excel"""
        raise NotImplementedError
    
    def format_stats_html(self, stats):
        """Formate les statistiques en HTML pour le PDF"""
        raise NotImplementedError


# Mixin spécifique pour ProductionExtrusion
class ProductionExtrusionExportMixin(ProductionExportMixin):
    
    def get_export_queryset(self, start_date, end_date):
        if start_date and end_date:
            return self.model.objects.filter(
                date_production__range=[start_date, end_date]
            ).select_related('zone', 'equipe')
        return self.model.objects.all().select_related('zone', 'equipe')
    
    def get_production_stats(self, queryset):
        if not queryset.exists():
            return None
        
        stats = queryset.aggregate(
            total_production=Sum('total_production_kg'),
            total_matiere=Sum('matiere_premiere_kg'),
            total_dechets=Sum('dechets_kg'),
            avg_rendement=Avg('rendement_pourcentage'),
            avg_taux_dechet=Avg('taux_dechet_pourcentage'),
            count=Count('id')
        )
        
        # Formater les valeurs
        for key, value in stats.items():
            if isinstance(value, Decimal):
                stats[key] = self.format_decimal(value, 2)
        
        return stats
    
    def format_stats_html(self, stats):
        return f"""
        <b>Statistiques de production:</b><br/>
        Nombre d'enregistrements: {stats['count']}<br/>
        Production totale: {stats['total_production']} kg<br/>
        Matière première utilisée: {stats['total_matiere']} kg<br/>
        Rendement moyen: {stats['avg_rendement']}%<br/>
        Déchets totaux: {stats['total_dechets']} kg<br/>
        Taux de déchets moyen: {stats['avg_taux_dechet']}%
        """
    
    def get_pdf_table_data(self, queryset):
        data = [['Date', 'Zone', 'Équipe', 'Chef Zone', 'Matière (kg)', 
                'Production (kg)', 'Rendement %', 'Déchets (kg)', 'Statut']]
        
        for prod in queryset:
            data.append([
                prod.date_production.strftime('%d/%m/%Y'),
                str(prod.zone),
                str(prod.equipe),
                prod.chef_zone,
                self.format_decimal(prod.matiere_premiere_kg),
                self.format_decimal(prod.total_production_kg) if prod.total_production_kg else "0.00",
                self.format_decimal(prod.rendement_pourcentage, 2) if prod.rendement_pourcentage else "0.00",
                self.format_decimal(prod.dechets_kg),
                "✓" if prod.valide else "✗"
            ])
        
        return data
    
    def create_excel_workbook(self, queryset, period):
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Extrusion {period}"
        
        # Styles
        styles = self.get_excel_styles()
        
        # En-têtes
        headers = [
            'Date', 'Zone', 'Équipe', 'Heure Début', 'Heure Fin', 
            'Chef Zone', 'Matière (kg)', 'Machines Actives', 'Machinistes',
            'Bobines (kg)', 'Finis (kg)', 'Semi-Finis (kg)', 'Déchets (kg)',
            'Total Production (kg)', 'Rendement %', 'Taux Déchet %', 
            'Production/Machine', 'Observations', 'Statut'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['center_alignment']
            cell.border = styles['thin_border']
        
        # Données
        for row_num, prod in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=prod.date_production)
            ws.cell(row=row_num, column=2, value=str(prod.zone))
            ws.cell(row=row_num, column=3, value=str(prod.equipe))
            ws.cell(row=row_num, column=4, value=prod.heure_debut.strftime('%H:%M') if prod.heure_debut else '')
            ws.cell(row=row_num, column=5, value=prod.heure_fin.strftime('%H:%M') if prod.heure_fin else '')
            ws.cell(row=row_num, column=6, value=prod.chef_zone)
            ws.cell(row=row_num, column=7, value=float(prod.matiere_premiere_kg))
            ws.cell(row=row_num, column=8, value=prod.nombre_machines_actives)
            ws.cell(row=row_num, column=9, value=prod.nombre_machinistes)
            ws.cell(row=row_num, column=10, value=float(prod.nombre_bobines_kg))
            ws.cell(row=row_num, column=11, value=float(prod.production_finis_kg))
            ws.cell(row=row_num, column=12, value=float(prod.production_semi_finis_kg))
            ws.cell(row=row_num, column=13, value=float(prod.dechets_kg))
            ws.cell(row=row_num, column=14, value=float(prod.total_production_kg) if prod.total_production_kg else 0)
            ws.cell(row=row_num, column=15, value=float(prod.rendement_pourcentage) if prod.rendement_pourcentage else 0)
            ws.cell(row=row_num, column=16, value=float(prod.taux_dechet_pourcentage) if prod.taux_dechet_pourcentage else 0)
            ws.cell(row=row_num, column=17, value=float(prod.production_par_machine) if prod.production_par_machine else 0)
            ws.cell(row=row_num, column=18, value=prod.observations[:100] if prod.observations else '')
            ws.cell(row=row_num, column=19, value="Validé" if prod.valide else "En attente")
        
        # Formater les colonnes numériques
        for col in ['G', 'J', 'K', 'L', 'M', 'N']:  # Colonnes avec valeurs numériques
            for row in range(2, len(queryset) + 2):
                cell = ws[f'{col}{row}']
                cell.number_format = '#,##0.00'
        
        for col in ['O', 'P', 'Q']:  # Pourcentages
            for row in range(2, len(queryset) + 2):
                cell = ws[f'{col}{row}']
                cell.number_format = '0.00%'
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ''
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb, ws
    
    def export_pdf_custom(self, request, start_date, end_date):
        # Utiliser la même logique que export_pdf_view mais avec des dates spécifiques
        period = f"personnalise_{start_date}_{end_date}"
        return self.export_pdf_view(request, period)
    
    def export_excel_custom(self, request, start_date, end_date):
        period = f"personnalise_{start_date}_{end_date}"
        return self.export_excel_view(request, period)


# Mixins similaires pour les autres modèles de production
class ProductionImprimerieExportMixin(ProductionExportMixin):
    # Implémentez les méthodes similaires à ProductionExtrusionExportMixin
    pass


class ProductionSoudureExportMixin(ProductionExportMixin):
    # Implémentez les méthodes similaires à ProductionExtrusionExportMixin
    pass


class ProductionRecyclageExportMixin(ProductionExportMixin):
    # Implémentez les méthodes similaires à ProductionExtrusionExportMixin
    pass